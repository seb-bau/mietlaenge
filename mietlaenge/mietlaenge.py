import json
import openpyxl
import requests

from dotenv import dotenv_values
from jsonmerge import Merger
from datetime import datetime


def openwowi_create_token(base_url, user, password, refresh_token=3600):
    # Token erstellen, der für die Dauer des Abrufs gültig ist.
    url = f"{base_url}/oauth2/token"

    payload = f"grant_type=password&" \
              f"username={user}&" \
              f"password={password}&" \
              f"refresh_token={refresh_token}"

    headers = {
        'Accept': 'text/plain',
        'Content-Type': 'application/x-www-form-urlencoded'
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    if response.status_code != 200:
        print(f"Fehler bei der Authentifizierung an OPEN WOWI. Status {response.status_code}:{response.text}")
        return False

    response_json = response.json()
    return response_json["access_token"]


def person_is_dead(base_url, token, api_key, person_id):
    # Prüfen, ob Person verstorben ist. Zweiter Rückgabewert ist das Personenobjekt
    # Dieser Schritt ist notwendig, da dem Contractor-Endpunkt leider die Information "DeathDate" fehlt.
    headers = {
        'Accept': 'text/plain',
        'Authorization': f'Bearer {token}'
    }
    url = f"{base_url}/openwowi/v1.0/PersonsRead/Person/{person_id}" \
          f"?apiKey={api_key}" \
          f"&includeCommunication=true" \
          f"&includeAddress=true"

    response = requests.request("GET", url, headers=headers)

    if response.status_code != 200:
        print(f"Fehler: {response.text}")
        exit()

    if response.json()["NaturalPerson"]["DeathDate"] is None:
        return False, response.json()
    else:
        return True, response.json()


def openwowi_get_all_license_agreements(base_url, token, api_key):
    # Alle Verträge aus Wowiport abrufen
    headers = {
        'Accept': 'text/plain',
        'Authorization': f'Bearer {token}'
    }
    complete_response = None
    merge_schema = {"mergeStrategy": "append"}
    merger = Merger(schema=merge_schema)

    t_offset = 0
    t_response_count = 100
    while t_response_count == 100:
        url = f"{base_url}/openwowi/v1.0/RentAccounting/LicenseAgreements" \
              f"?apiKey={api_key}" \
              f"&offset={t_offset}" \
              f"&limit=100"

        response = requests.request("GET", url, headers=headers)

        if response.status_code != 200:
            print(response.text)
            return None

        if complete_response is None:
            complete_response = response.json()
        else:
            complete_response = merger.merge(complete_response, response.json())

        t_response_count = len(response.json())
        t_total_contracts = len(complete_response)
        t_offset += 100
        print(f"LA total: {t_total_contracts}")
    return complete_response


def openwowi_get_all_contractors(base_url, token, api_key):
    # Alle Vertragsnehmer aus Wowiport abrufen
    headers = {
        'Accept': 'text/plain',
        'Authorization': f'Bearer {token}'
    }
    complete_response = None
    merge_schema = {"mergeStrategy": "append"}
    merger = Merger(schema=merge_schema)

    t_offset = 0
    t_response_count = 100
    while t_response_count == 100:
        url = f"{base_url}/openwowi/v1.1/RentAccountingPersonDetails/Contractors" \
              f"?apiKey={api_key}" \
              f"&offset={t_offset}" \
              f"&limit=100" \
              f"&includeMainAddress=true" \
              f"&includeMainCommunication=true" \
              f"&includePersonAddresses=true" \
              f"&includePersonCommunications=true" \
              f"&includePersonBankAccounts=true"

        response = requests.request("GET", url, headers=headers)

        if response.status_code != 200:
            print(f"Fehler in openwowi_get_all_contractors: {response.status_code} - {response.text}")
            return None

        if complete_response is None:
            complete_response = response.json()
        else:
            complete_response = merger.merge(complete_response, response.json())

        t_response_count = len(response.json())
        t_total_contractors = len(complete_response)
        t_offset += 100
        print(f"Total: {t_total_contractors}")
    return complete_response


def contractors_of_contract(pcontractors, contract_id):
    # Nur 1. und 2. Vertragsnehmer, abweichender Vertragsnehmer und Untermieter.
    # Abrufbar mit  /openwowi/v1.0/RentAccountingPersonDetails/ContractorTypes
    valid_types = [1, 2, 3, 1003]
    ret_contractors = []
    for contractor in pcontractors:
        if contractor["LicenseAgreementId"] == contract_id:
            if contractor["ContractorType"]["Id"] in valid_types:
                if contractor["Person"]["IsNaturalPerson"]:
                    ret_contractors.append(contractor)
    return ret_contractors


def get_duration_days(spans):
    # Dauer der Zeitspannen berechnen und Überlappungen abziehen
    duration_days = 0
    overlaps = []
    for i, span_i in enumerate(spans):
        span_duration = (span_i[1] - span_i[0]).days + 1
        duration_days += span_duration
        for span_j in spans[i + 1:]:
            overlap_start = max(span_i[0], span_j[0])
            overlap_end = min(span_i[1], span_j[1])
            overlap_duration = (overlap_end - overlap_start).days + 1
            if overlap_duration > 0:
                overlap = (overlap_start, overlap_end, overlap_duration)
                if overlap not in overlaps:
                    overlaps.append(overlap)

    # Dauer der Überlappungen abziehen
    for overlap in overlaps:
        duration_days -= overlap[2]

    # Ausgabe der Gesamtdauer
    return duration_days


def liste_mitlaenge():
    settings = dotenv_values(".env")
    wowi_host = settings.get("wowi_url")
    wowi_user = settings.get("wowi_user")
    wowi_pass = settings.get("wowi_pass")
    wowi_key = settings.get("wowi_api_key")

    token = openwowi_create_token(wowi_host, wowi_user, wowi_pass)

    use_cache = True
    cache_file_wowi_nv = "wowi_nv.json"
    cache_file_wowi_contractors = "wowi_contractors.json"
    if use_cache:
        with open(cache_file_wowi_nv) as json_file:
            wowi_vertraege = json.load(json_file)

        with open(cache_file_wowi_contractors) as json_file:
            wowi_contractors = json.load(json_file)
    else:
        wowi_vertraege = openwowi_get_all_license_agreements(wowi_host, token, wowi_key)
        wowi_la_json_string = json.dumps(wowi_vertraege)
        with open(cache_file_wowi_nv, 'w') as outfile:
            outfile.write(wowi_la_json_string)

        wowi_contractors = openwowi_get_all_contractors(wowi_host, token, wowi_key)
        wowi_contractors_json_string = json.dumps(wowi_contractors)
        with open(cache_file_wowi_contractors, 'w') as outfile:
            outfile.write(wowi_contractors_json_string)

    # WIE-Filter vorbereiten
    wie_filter = settings.get("wie_filter", "")
    wie_starts = None
    if len(wie_filter.strip()) > 0:
        wie_starts = wie_filter.strip().split(";")
    mieter = {}
    contract_counter = 0
    for vertrag in wowi_vertraege:
        contract_counter += 1
        if wie_starts is not None:
            wie_valid = False
            for start_value in wie_starts:
                if str(vertrag["UseUnit"]["EconomicUnit"]).startswith(start_value):
                    wie_valid = True
                    break
        else:
            wie_valid = True

        if not wie_valid:
            continue

        use_unit_number_parts = vertrag["UseUnit"]["UseUnitNumber"].split(".")
        building_num = int(use_unit_number_parts[1])

        obj_filter = int(settings.get("obj_filter", 0))
        if obj_filter > 0:
            if building_num >= obj_filter:
                continue

        contractors = contractors_of_contract(wowi_contractors, vertrag["Id"])

        now = datetime.now()

        if len(contractors) > 0:
            for vertragsnehmer in contractors:
                currently_valid = False
                stamp_start = vertragsnehmer["ContractualUseValidFrom"]
                stamp_stop = vertragsnehmer["ContractualUseValidTo"]
                start_dt = datetime.strptime(stamp_start, "%Y-%m-%d")
                if stamp_stop is not None:
                    stop_dt = datetime.strptime(stamp_stop, "%Y-%m-%d")
                    if stop_dt > now:
                        currently_valid = True
                        stop_dt = now
                else:
                    if vertragsnehmer["EndOfContract"] is not None:
                        stop_dt = datetime.strptime(vertragsnehmer["EndOfContract"], "%Y-%m-%d")
                        if stop_dt > now:
                            currently_valid = True
                            stop_dt = now
                    else:
                        currently_valid = True
                        stop_dt = now

                person_num = vertragsnehmer["Person"]["IdNum"]
                if person_num in mieter.keys():
                    mieter_dict = mieter.get(person_num)
                    cspans = mieter_dict["spans"]
                    cspans.append((start_dt, stop_dt))
                    mieter_dict["spans"] = cspans
                    mieter_dict["tage_mieter"] = get_duration_days(cspans)
                    if currently_valid:
                        mieter_dict["contractor"] = vertragsnehmer
                    mieter[person_num] = mieter_dict
                else:
                    nspans = [(start_dt, stop_dt)]
                    tdelta = get_duration_days(nspans)
                    tdict = {
                        "tage_mieter": tdelta,
                        "spans": nspans
                    }
                    if currently_valid:
                        tdict["contractor"] = vertragsnehmer
                    mieter[person_num] = tdict

    mieter_output = {}
    for entrynum in mieter:
        entry = mieter.get(entrynum)
        if "contractor" not in entry.keys():
            # Kein aktueller MV gefunden
            continue

        entry_cont = entry.get("contractor")
        if entry_cont["Person"]["ValidTo"] is not None:
            # Person gestorben
            continue

        mieter_tage = entry["tage_mieter"]
        mieter_jahre = mieter_tage / 365

        if mieter_jahre < 40:
            continue
        is_dead, person_object = person_is_dead(wowi_host, token, wowi_key, entry_cont['Person']['Id'])
        entry["person_object"] = person_object
        if is_dead:
            continue
        print(f"Mieter {entry_cont['Person']['IdNum']} Jahre: {mieter_jahre}")
        mieter_output[entry_cont["Person"]["IdNum"]] = entry

    output_file = "output.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Vorname"
    sheet.cell(row=1, column=3).value = "Titel"
    sheet.cell(row=1, column=4).value = "Anrede"
    sheet.cell(row=1, column=5).value = "Straße"
    sheet.cell(row=1, column=6).value = "PLZ"
    sheet.cell(row=1, column=7).value = "Ort"
    sheet.cell(row=1, column=8).value = "Geburtstag"
    sheet.cell(row=1, column=9).value = "Personennummer"
    sheet.cell(row=1, column=10).value = "Jahre Mieter"
    rowcount = 2
    for line in mieter_output:
        line_entry = mieter_output.get(line)
        main_address = None
        if line_entry["person_object"]["Addresses"] is None:
            continue
        for address in line_entry["person_object"]["Addresses"]:
            if address["MainAddress"]:
                main_address = address
                break

        bdate = line_entry["person_object"]["NaturalPerson"]["BirthDate"]
        bdate_str = bdate.split("T")[0]

        geschlecht = line_entry["person_object"]["NaturalPerson"]["Gender"]["Name"]
        titel = line_entry["person_object"]["NaturalPerson"]["Title"]
        nachname = line_entry["person_object"]["NaturalPerson"]["LastName"]
        vorname = line_entry["person_object"]["NaturalPerson"]["FirstName"]
        if titel is not None and len(titel.strip()) > 0:
            titelstring = f"{titel} "
        else:
            titelstring = ""
        if geschlecht == "männlich":
            anrede = f"Sehr geehrter Herr {titelstring}{nachname},"
        elif geschlecht == "weiblich":
            anrede = f"Sehr geehrte Frau {titelstring}{nachname},"
        else:
            anrede = f"Guten Tag {titelstring}{vorname} {nachname},"

        sheet.cell(row=rowcount, column=1).value = nachname
        sheet.cell(row=rowcount, column=2).value = vorname
        sheet.cell(row=rowcount, column=3).value = titel
        sheet.cell(row=rowcount, column=4).value = anrede
        sheet.cell(row=rowcount, column=5).value = main_address.get("StreetComplete")
        sheet.cell(row=rowcount, column=6).value = main_address.get("Zip")
        sheet.cell(row=rowcount, column=7).value = main_address.get("Town")
        sheet.cell(row=rowcount, column=8).value = bdate_str
        sheet.cell(row=rowcount, column=9).value = line_entry["contractor"]["Person"]["IdNum"]
        sheet.cell(row=rowcount, column=10).value = str(round(line_entry["tage_mieter"] / 365))
        rowcount += 1

    wb.save(output_file)


liste_mitlaenge()
